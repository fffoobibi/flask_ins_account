from flask import request, jsonify
from functools import wraps
from datetime import datetime, timedelta
import setting


# 定义装饰器函数
def validate_token(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        # 从请求中获取 token
        token = request.headers.get("check-token")
        # 检查 token 的有效性
        if token == "1C168F31197B4C4F":
            return f(*args, **kwargs)
        else:
            return jsonify({"error": "Invalid token"}), 401

    return wrapper


def format_datetime_fields(
    dict_records: list[dict],
    datetime_fields: list[str],
    datetime_format: str = "%Y-%m-%d %H:%M:%S",
):
    """
    格式化字典中的日期时间字段。
    """
    for record in dict_records:
        for field in datetime_fields:
            if record[field]:
                record[field] = record[field].strftime(datetime_format)
    return dict_records

def get_dt_range(
        days: int,
        end: datetime = None,
) -> tuple[datetime, datetime]:
    """
    获取指定日期范围

    Args:
        days (int): 间隔天数
        end (datetime, optional): 结束日期. Defaults to None.

    Returns:
        tuple[datetime, datetime]: (start_date, end_date)
    """

    if end is None:
        end = datetime.now().strftime("%Y-%m-%d 23:59:59")
        end_date = datetime.strptime(end, "%Y-%m-%d %H:%M:%S")
    else:
        end_date = end
    start_date = (end_date - timedelta(days=days)).strftime("%Y-%m-%d 00:00:00")
    start_date = datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")
    return start_date, end_date

def save_to_excel(
        data: list[dict],
        headers: dict[str, str],
        file_name: str,
        column_widths: dict[str, int] = None,
        header_height=38,
        row_height=40,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active

    # 设置表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")  # 表头居中对齐

    # 写入表头
    ws.row_dimensions[1].height = header_height
    for col_num, (field, title) in enumerate(headers.items(), start=1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment  # 应用居中对齐
        if column_widths and field in column_widths:
            ws.column_dimensions[get_column_letter(col_num)].width = column_widths[
                field
            ]

    # 写入数据
    for row_idx, record in enumerate(data, start=2):
        ws.row_dimensions[row_idx].height = row_height
        for col_num, field in enumerate(headers.keys(), start=1):
            # 处理普通字段
            value = record.get(field, "")
            ws.cell(row=row_idx, column=col_num, value=value)

    # 保存Excel文件
    fp = "./uploads/" + file_name
    wb.save(fp)
    return setting.upload_path + file_name